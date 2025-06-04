import argparse
import os
import csv
from datetime import datetime
from openpyxl import Workbook

CONFIG_DEFAULT = "output"

# List of replacements to apply to sheet names
SHEET_NAME_REPLACEMENTS = [
    ("_filtered_users", ""),
    ("users_", "")
]

def parse_args():
    parser = argparse.ArgumentParser(description="Uses all CSV in the folder to create a XLSX. You might change the folder it uses for input and output")
    parser.add_argument(
        "--output",
        type=str,
        default=CONFIG_DEFAULT,
        help=f"Path to the where the CSV are stored and where the XLSX will be written to (by default {CONFIG_DEFAULT})"
    )
    return parser.parse_args()


def clean_sheet_name(name):
    for old, new in SHEET_NAME_REPLACEMENTS:
        name = name.replace(old, new)
    return name[:31]  # Excel sheet name limit

def csv_to_worksheet(wb, csv_path, sheet_name):
    ws = wb.create_sheet(title=sheet_name)

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row_index, row in enumerate(reader, 1):
            for col_index, value in enumerate(row, 1):
                ws.cell(row=row_index, column=col_index, value=value)

def combine_all_csvs(output_dir):
    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_file = os.path.join(output_dir, f"combined_output_{timestamp}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    numfiles = 0
    for file in sorted(os.listdir(output_dir)):
        if file.endswith(".csv"):
            numfiles = numfiles + 1
            path = os.path.join(output_dir, file)
            raw_name = os.path.splitext(file)[0]
            sheet_name = clean_sheet_name(raw_name)
            print(f"[+] Adding sheet: {sheet_name}")
            csv_to_worksheet(wb, path, sheet_name)

    if numfiles > 0:
        wb.save(output_file)
        print(f"[✓] Excel file created: {output_file}")
    else:
        print(f"[✗] No input files available in '{output_dir}' folder")

if __name__ == "__main__":
    args = parse_args()
    combine_all_csvs(args.output)
